import pandas as pd
import os
import glob # In your local windows machine, .glob is used to find pathnames matching a specified pattern.
import chardet
import csv
from datetime import datetime
import re # to extract the version number from a file name
import numpy as np #use this to make a conditional value assignments to the new accounts as 1 if new and 0 if not new
import shutil # use this to copy files
from openpyxl import load_workbook # use this to load an existing Excel workbook and add a sheet
from pandas.core.frame import DataFrame # use this to convert a dataframe to a list of records that can be added to a excel sheet 

#Setup File locations
#=================================== Deduplicated accounts data, this file contains the list of new and historical accounts that has been deduplicated i.e., grouped into clusters

# Define the base directory and the file prefix
base_directory = r"C:\Users\beste\OneDrive - Qral Group\01 Narcan\12 Ops"
file_prefix = 'P02_Dedupped_Results__'

# Use glob to find all files that match the prefix in all subdirectories of the base directory. the ** means that it will search all subdirectories of the base directory. The *.csv means that it will search for files with the .csv extension. So, it will look for the file prefix and the .csv extension in all subdirectories of the base directory, regardless of how many levels deep they are.
files = glob.glob(os.path.join(base_directory, '**', file_prefix + '*.csv'), recursive=True) #The recursive=True argument means that it will search all subdirectories of the subdirectories, and so on, regardless of how many levels deep they are.

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix}' in directory '{base_directory}'")
    exit(1)

# Sort the files by date, and select the most recent one
latest_Deduped_Result_file = max(files, key=os.path.getctime)

print(f"Using file '{latest_Deduped_Result_file}'") # this will print the most recent file name that will be used for the deduplication process

# Get the parent directory of the latest file
input2_base_path = os.path.dirname(latest_Deduped_Result_file)

print(f"Using base path '{input2_base_path}'") # this will print the base path that will be used for loading files and saving the results of the deduplication process


#=================================== Salesforce accounts data, this file contains the most recent accounts in salesforce to be used for reference when deciding which account are new accounts
file_prefix2 = 'SFDC Accounts Extract'

# Use glob to find all files that match the prefix in all subdirectories of the base directory
files = glob.glob(os.path.join(base_directory, '**', file_prefix2 + '*.xlsx'), recursive=True)

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix2}' in directory '{base_directory}'")
    exit(1)

# Sort the files by date, and select the most recent one
latest_SFDC_extract_file = max(files, key=os.path.getctime)

print(f"Using file '{latest_SFDC_extract_file}'") # this will print the full path of the most recent file that will be used for checking new accounts against the logbook

# Get the parent directory of the latest file
input3_base_path = os.path.dirname(latest_SFDC_extract_file)

print(f"Using base path '{input3_base_path}'") # this will print the base path that will be used for loading files and saving the results of the analysis of new accounts

#latest_file = os.path.join(input2_base_path, 'Mastering Full Extract 20231218_114447.xlsx')

sfdc_extract_filename = os.path.basename(latest_SFDC_extract_file) # use this to print the name of the file that will be used for checking new accounts against the logbook



# Function that automatically detects the character encoding needed to read file
def detect_file_encoding(file_path):
    with open(file_path, 'rb') as file:
        result = chardet.detect(file.read())
        return result['encoding']

# use chardet to detect the encoding of the csv file
file_encoding = detect_file_encoding(latest_Deduped_Result_file)

#Read in raw_input1_file csv file as a dataframe. This is the dedupped SpecDistr output file
deduped_input_data= pd.read_csv(latest_Deduped_Result_file, encoding=file_encoding, keep_default_na=False) #keep_default_na=False prevents pandas from reading blank values as NaN

#print(deduped_input_data['MDM ID'].unique())

#Rename 'MDM ID' to 'Final MDM ID'
#deduped_input_data = deduped_input_data.rename(columns={'MDM ID': 'Final MDM ID'})

print(deduped_input_data['Final MDM ID'].unique())

#=================================== Define the output file path and filename



# Get the filename from the input file path
input_filename = os.path.basename(latest_Deduped_Result_file)

# Add the prefix and suffix to the filename
output_filename = f"MDM_ID_assigned_{input_filename.replace('.csv', '')}.csv" # We use the replace method to remove the '.csv' extension from the filename. Then, we add the prefix and suffix specified in the f string

# Add the prefix and suffix to the filename
output2_filename = f"P03a_MDM_ID_assigned_net_new_accounts_only_{input_filename.replace('.csv', '')}.csv" # We use the replace method to remove the '.csv' extension from the filename. Then, we add the prefix and suffix specified in the f string


# Create the full path for the output file
output_file = os.path.join(input2_base_path, output_filename) # This will contains all accounts, including historical accounts

# Create the full path for the output file
output_file2 = os.path.join(input2_base_path, output2_filename) # Use this to save the net new account separately from the full results. This is used as the input for fuzzy match in recordlink.py

#=================================== Historical Accounts Data: Customer Mastering LogBook file

# Define the base directory and the file prefix
base_directory = r"C:\Users\beste\OneDrive - Qral Group\01 Narcan\08 Mastering"
file_prefix = 'Customer Mastering LogBook v'

# Use glob to find all files that match the prefix in all subdirectories of the base directory
files = glob.glob(os.path.join(base_directory, '**', file_prefix + '*.xlsx'), recursive=True)

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix}' in directory '{base_directory}'")
    exit(1)

# Define a function to extract the version number from a file name
def get_version_number(filename):
    match = re.search(r'v(\d+\.\d+)', filename)
    if match:
        return float(match.group(1))
    else:
        return 0

# Sort the files by version number, and select the one with the highest version number
latest_mastering_logbook_file = max(files, key=get_version_number)

print(f"Using file '{latest_mastering_logbook_file}'") # this will print the full path of the most recent file that will be used for checking new accounts against the logbook

# Get the parent directory of the latest file
input1_base_path = os.path.dirname(latest_mastering_logbook_file)

print(f"Using base path '{input1_base_path}'") # this will print the base path that will be used for loading the logbook file

#raw_input1_file = os.path.join(input1_base_path, 'Customer Mastering LogBook v1.12 - Pre1218.xlsx') # Set path containing the file with previously analyzed accounts

logbook_master_data_filename = os.path.basename(latest_mastering_logbook_file) #extract filename of logbook from path for use in print statements

#raw_input2_file = os.path.join(input1_base_path, 'Customer Mastering LogBook v1.12 - Pre1218.xlsx') # Set path containing the file with previously analyzed accounts with Final MDM ID already assigned

#logbook_master_data_filename = os.path.basename(raw_input2_file) #extract filename of logbook from path for use in print statements

#Read in raw_input1_file file as a dataframe. This is the master logbook historical account list from a EXCEL workbook against which we check the new accounts
logbook_master_data = pd.read_excel(latest_mastering_logbook_file, sheet_name='AccountMaster')


#=================================== Set up Final MDM ID for new accounts: starting from the max Final MDM ID in the historical accounts data, increment by 1 and assign to each new account


# Convert 'Final MDM ID' to string
logbook_master_data['Final MDM ID'] = logbook_master_data['Final MDM ID'].astype(str)


# Filter out rows where 'Final MDM ID' is blank or contains the string 'pending'.
# The ~ operator is the bitwise NOT operator. It returns True if the operand is False and False if the operand is True. In this case, it returns True for rows where 'Final MDM ID' is not blank and does not contain the string 'pending'
filtered_logbook_master_data = logbook_master_data[(logbook_master_data['Final MDM ID'] != '') & (~logbook_master_data['Final MDM ID'].str.contains('pending'))]

# Extract the numeric part of 'Final MDM ID', convert to float. We convert to a float because the max value of a string is not the same as the max value of a number. So, we need to convert to a number to find the max value. Final MDM ID Numeric is a new column that we are creating in the filtered_logbook_master_data dataframe
filtered_logbook_master_data['Final MDM ID Numeric'] = filtered_logbook_master_data['Final MDM ID'].str.replace('EMUSHCO', '').astype(float) # the str.replace method replaces the string 'EMUSHCO' with an empty string, which effectively removes it from the string. The astype method converts the resulting string to a float

# Replace NaN values with a default value (e.g., 0), then convert to integer. We convert to an integer because we want to increment the max value by 1
filtered_logbook_master_data['Final MDM ID Numeric'] = filtered_logbook_master_data['Final MDM ID Numeric'].fillna(0).astype(int) # the fillna method replaces NaN values with the value specified in the argument. In this case, we replace NaN values with 0

# Find the max of 'Final MDM ID Numeric'
max_mcm_id_numeric = filtered_logbook_master_data['Final MDM ID Numeric'].max()

# Construct the max 'Final MDM ID' using the max numeric part
max_mcm_id = f"EMUSHCO{max_mcm_id_numeric:07d}" # the :07d part means that the number should be formatted as a 7-digit integer with leading zeros

# Print the max Final MDM ID value in the filtered_logbook_master_data dataframe as an f string to check that the max Final MDM ID value is correct
print(f"Max Final MDM ID value in historical {logbook_master_data_filename} is : {max_mcm_id}")

# Extract the numeric part of the max Final MDM ID and increment it by 1
if pd.isnull(max_mcm_id):
    next_id = 1
else:
    next_id = int(max_mcm_id.replace("EMUSHCO", "")) + 1
    

 # Add a print statement to check the start_id value as an f string
print(f"The starting Final MDM ID value for new accounts is: {next_id}")

# Convert 'Final MDM ID' to string in deduped_input_data
deduped_input_data['Final MDM ID'] = deduped_input_data['Final MDM ID'].astype(str)

# Identify the rows in the input data where the 'Final MDM ID' column contains the string 'pending'. If there are no rows with 'pending' in the 'Final MDM ID' column, then we check if there are any rows with blank 'Final MDM ID' values
pending_rows = deduped_input_data['Final MDM ID'].str.contains('pending') | deduped_input_data['Final MDM ID'].isnull() | (deduped_input_data['Final MDM ID'] == '') # the | operator is the bitwise OR operator. It returns True if either of the operands is True
print(f"Number of pending acounts: {pending_rows.sum()}")

#Print pre-exisintg Final MDM IDs that are not pending and not blank as a QC check. These should be the same as the number of rows in that have an Final MDM ID assigned in the historical accounts data
print(f" Number of historical non pending accounts: {deduped_input_data[~deduped_input_data['Final MDM ID'].str.contains('pending') & deduped_input_data['Final MDM ID'].notnull() & (deduped_input_data['Final MDM ID'] != '')].shape[0]}") # this line prints the number of rows that do not contain 'pending' in the 'Final MDM ID' column and are not blank

# For each of these rows, check if there are any other rows with the same 'Cluster ID' that already have an 'Final MDM ID' assigned
for index, row in deduped_input_data[pending_rows].iterrows():
    cluster_id = row['Cluster ID']
    #print(f"Current index: {index}, Cluster ID: {cluster_id}, Final MDM ID: {row['Final MDM ID']}")
    #existing_mdm_id_rows = deduped_input_data[(deduped_input_data['Cluster ID'] == cluster_id) & (~deduped_input_data['Final MDM ID'].str.contains('pending'))& (deduped_input_data['Final MDM ID'].notnull())] # Use this if you want to check for any existing Final MDM ID, not just the blank ones. This is useful if you want to re-run the script on a file that already has some Final MDM IDs assigned and includes pending rows
    existing_mdm_id_rows = deduped_input_data[(deduped_input_data['Cluster ID'] == cluster_id) & (~deduped_input_data['Final MDM ID'].isnull()) & (deduped_input_data['Final MDM ID'] != '')]
    
    # If there are, assign the existing 'Final MDM ID' to the 'pending' row
    if not existing_mdm_id_rows.empty:
        #print(f"Existing Final MDM ID assigned: {existing_mdm_id_rows['Final MDM ID'].iloc[0]}") # this prints the first value in the 'Final MDM ID' column of the existing_mdm_id_rows dataframe. It is used to check that the correct Final MDM ID is being assigned
        deduped_input_data.loc[index, 'Final MDM ID'] = existing_mdm_id_rows['Final MDM ID'].iloc[0]
    # If there aren't, assign a new 'Final MDM ID' that is 1 greater than the maximum 'Final MDM ID' assigned so far
    else:
        deduped_input_data.loc[index, 'Final MDM ID'] = f"EMUSHCO{next_id:07d}"
        #print(f"New Final MDM ID assigned: EMUSHCO{next_id:07d}") # this prints the new Final MDM ID that is being assigned. It is used to check that the correct Final MDM ID is being assigned
        next_id += 1  # Increment the next_id for the next assignment
  
# Print the max Final MDM ID value in the deduped_input_data dataframe as an f string to check that the max Final MDM ID value is correct
print(f"Max Final MDM ID value in deduped_input_data is: {deduped_input_data['Final MDM ID'].max()}")


###================================== QC Check that the number of unique 'Final MDM ID' values is equal to the number of unique 'Cluster ID' values


# Get the number of unique 'Cluster ID'
num_unique_cluster_id = deduped_input_data['Cluster ID'].nunique()

# Get the number of unique 'Final MDM ID'
num_unique_mdm_id = deduped_input_data['Final MDM ID'].nunique()

# Print the number of unique 'Cluster ID' and 'Final MDM ID'
print(f"Number of unique Cluster ID: {num_unique_cluster_id}")
print(f"Number of unique Final MDM ID: {num_unique_mdm_id}")

# Get a boolean series where True indicates the 'Final MDM ID' is duplicated
duplicated_mdm_id = deduped_input_data.duplicated('Final MDM ID', keep=False) # the duplicated method returns a boolean series where True indicates the value is duplicated. The keep=False argument indicates that all duplicated values should be marked as True

# Create a new DataFrame that only includes the rows with duplicated 'Final MDM ID' values
duplicated_mdm_id_df = deduped_input_data[duplicated_mdm_id]

# Sort the DataFrame by 'Final MDM ID'
sorted_duplicated_mdm_id_df = duplicated_mdm_id_df.sort_values('Final MDM ID')

# Print the sorted DataFrame
print(sorted_duplicated_mdm_id_df)

#print an f string that QC checks are complete
print(f"QC checks are complete")

###==============Create a new column to flag New accounts, Reorder the columns in the deduped_input_data dataframe to bring MDMD ID to the front


# Create 'Net New Accounts' column
deduped_input_data['Net New Accounts'] = np.where(deduped_input_data['Final MDM ID'] > max_mcm_id, 1, 0)

# Get a list of all columns
cols = list(deduped_input_data.columns)

# Remove 'Final MDM ID' and 'Net New Accounts' from the list
cols.remove('Final MDM ID')
cols.remove('Net New Accounts')

# Create a new list with 'Final MDM ID' and 'Net New Accounts' at the start
cols = ['Final MDM ID', 'Net New Accounts'] + cols

# Reindex the DataFrame with the new column order
deduped_input_data = deduped_input_data[cols]




###==================================Save a log of Final MDM IDs generated and time stamp

# Specify the full path to the file
mdm_id_log_file_path = r'C:\Users\beste\OneDrive - Qral Group\01 Narcan\08 Mastering\00 Logbook\MDM_ID_creation_log.csv'


# Calculate the number of Final MDM IDs generated
num_mdm_ids = deduped_input_data['Final MDM ID'].count()

# Calculate the maximum Final MDM ID generated
max_new_mdm_id = deduped_input_data['Final MDM ID'].max()

# Filter the DataFrame to include only rows where 'Final MDM ID' is greater than max_mcm_id
new_mdm_ids = deduped_input_data[deduped_input_data['Final MDM ID'] > max_mcm_id]

# Calculate the number of unique 'Final MDM ID' in the filtered DataFrame
num_new_mdm_ids = new_mdm_ids['Final MDM ID'].nunique()

print(f"Number of new Final MDM IDs created: {num_new_mdm_ids}")

###================================== Save only the new_mdm_ids to a CSV file

# Load the SFDC Accounts Extract file into a DataFrame
sfdc_data = pd.read_excel(latest_SFDC_extract_file)

# Get a list of unique 'ShipTo Number' values
shipto_numbers = sfdc_data['ShipTo Number'].unique()

# Filter new_mdm_ids to include only rows where 'ShipTo Code' is not in shipto_numbers
new_only_mdm_ids = new_mdm_ids[~new_mdm_ids['ShipTo Code'].isin(shipto_numbers)]

# Define a dictionary mapping the old column names to the new column names
column_names = {
    'Cluster ID': 'Dedupe_Cluster ID',
    'confidence_score': 'Dedupe_confidence_score',
    'Address': 'ShippingStreet',
    'City': 'ShippingCity',
    'State': 'ShippingState',
    'Zip': 'ShippingPostalCode'
}

# Rename the columns
new_only_mdm_ids = new_only_mdm_ids.rename(columns=column_names)

#Save to csv file
new_only_mdm_ids.to_csv(output_file2, index=False) # This is used as one of the inputs to fuzzy match to salesforce in record_linkage.py

print(f"Records with new MDM IDs saved to CSV file. Number of records in the new account file is: {len(new_only_mdm_ids)}")

# Filter new_mdm_ids to include only rows where 'ShipTo Code' is in shipto_numbers
existing_mdm_ids = new_mdm_ids[new_mdm_ids['ShipTo Code'].isin(shipto_numbers)]

print(f"Number of records among new accounts with ShipTo Number found in Saleforce extract is: {len(existing_mdm_ids)}")

###==================================Save the full detailed Results of this analysis data to EXCEL, including historical accounts

# Define the output filename and path for the Excel file
output_excel_filename = f"P03b_MDM_ID_assigned_All_accounts_{input_filename.replace('.csv', '.xlsx')}"
output_excel_file = os.path.join(input2_base_path, output_excel_filename)

# Create a Pandas Excel writer using XlsxWriter as the engine
with pd.ExcelWriter(output_excel_file, engine='xlsxwriter') as writer:
    # Write each DataFrame to a different worksheet
    deduped_input_data.to_excel(writer, sheet_name='All Accounts', index=False)
    existing_mdm_ids.to_excel(writer, sheet_name='New Acc ShipTo in Salesforce', index=False)
    new_mdm_ids.to_excel(writer, sheet_name='Net New Accounts', index=False)

print(f"Data saved to Excel file '{output_excel_filename}' with 'All Accounts', 'Existing Accounts', and 'New Accounts' as separate sheets.")

###==================================Create a copy of the Mastering Analysis needed for ingestion of new account updates, to EXCEL
# Define the base directory and the file prefix
base_directory = r"C:\Users\beste\OneDrive - Qral Group\01 Narcan\08 Mastering"
file_prefix = 'Mastering All Sources'

# Use glob to find all files that match the prefix in all subdirectories of the base directory
files = glob.glob(os.path.join(base_directory, '**', file_prefix + '*.xlsx'), recursive=True)

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix}' in directory '{base_directory}'")
    exit(1)

# Define a function to extract the version number from a file name
def get_version_number(filename):
    match = re.search(r'v(\d+\.\d+)', filename)
    if match:
        return float(match.group(1))
    else:
        return 0

# Sort the files by version number, and select the one with the highest version number
latest_Mastering_All_Sources_Analysis_file = max(files, key=get_version_number)

# Extract the version number from the latest file
latest_version = get_version_number(latest_Mastering_All_Sources_Analysis_file)

# Increment the version number by 0.01
new_version = latest_version + 0.01

# Replace the old version number in the file name with the new version number
new_file = re.sub(r'v\d+\.\d+', f'v{new_version:.2f}', latest_Mastering_All_Sources_Analysis_file)

# Copy the latest file to the new file
shutil.copy(latest_Mastering_All_Sources_Analysis_file, new_file)

print(f"Copied file '{latest_Mastering_All_Sources_Analysis_file}' to '{new_file}'")

# Your existing code for finding the latest file and creating a copy goes here...

# Load the existing workbook
book = load_workbook(new_file)

# Create a new sheet in the workbook for 'new_mdm_ids'
new_sheet1 = book.create_sheet('1a. Net New Accounts input')

# Write the column headers to the new sheet
new_sheet1.append(new_mdm_ids.columns.tolist())

# Write the DataFrame to the new sheet
for r in new_mdm_ids.to_dict(orient='records'):  # Convert DataFrame to list of dicts
    new_sheet1.append(list(r.values()))  # Convert each dict to a list of values

# Create a new sheet in the workbook for 'sfdc_data'
new_sheet2 = book.create_sheet('1c. SFDC Acc Extract')

# Write the column headers to the new sheet
new_sheet2.append(sfdc_data.columns.tolist())

# Write the DataFrame to the new sheet
for r in sfdc_data.to_dict(orient='records'):  # Convert DataFrame to list of dicts
    new_sheet2.append(list(r.values()))  # Convert each dict to a list of values

# Save the workbook
book.save(new_file)

print(f"Added new sheets to '{new_file}'")


### ========================================Save the log of new account created to the log file

# Get the current timestamp
timestamp = datetime.now()

# Open the CSV file in append mode
with open(mdm_id_log_file_path, 'a', newline='') as csv_file:
    writer = csv.writer(csv_file)
    
    # If the file is empty, write the column names
    if os.stat('MDM_ID_creation_log.csv').st_size == 0:
        writer.writerow(["Date", "Number of Final MDM IDs Created", "Max Final MDM ID"])
    
    # Write the timestamp, the number of Final MDM IDs generated, and the maximum Final MDM ID generated to the CSV file
    writer.writerow([timestamp, num_new_mdm_ids, max_new_mdm_id])