import pandas as pd
from fuzzywuzzy import fuzz
import os
from datetime import datetime
import glob # In your local windows machine, .glob is used to find pathnames matching a specified pattern.
from openpyxl import load_workbook # use this to load an existing Excel workbook and add a sheet
from pandas.core.frame import DataFrame # use this to convert a dataframe to a list of records that can be added to a excel sheet 
import re # to extract the version number from a file name

###=================================== Net new accounts data, this file contains the list of new accounts that is to be mapped to records in salesforce

# Define the base directory and the file prefix
base_directory = r"C:\Users\beste\OneDrive - Qral Group\01 Narcan\12 Ops"
file_prefix = 'P04a_Customer_mastering_output'

# Use glob to find all files that match the prefix in all subdirectories of the base directory
files = glob.glob(os.path.join(base_directory, '**', file_prefix + '*.xlsx'), recursive=True)

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix}' in directory '{base_directory}'")
    exit(1)

# Sort the files by date, and select the most recent one
latest_Fuzzy_Result_file = max(files, key=os.path.getctime)

print(f"Using file '{latest_Fuzzy_Result_file}'") # this will print the full path of the most recent file that will be used for checking new accounts against the logbook

# Get the parent directory of the latest file
base_path = os.path.dirname(latest_Fuzzy_Result_file)

print(f"Using base path '{base_path}'") # this will print the base path that will be used for loading files and saving the results of the analysis of new accounts

#latest_file = os.path.join(input2_base_path, 'Mastering Full Extract 20231218_114447.xlsx')

latest_fuzzy_result_filename = os.path.basename(latest_Fuzzy_Result_file) # use this to print the name of the file that will be used for checking new accounts against the logbook


###=============================================================================

# read file in pandas
Processed_input_subdirectory = "mastering"


# create variable that stores today's data and time by # Generate the current date and time and then convrt it into string to use as a file name suffix in 'YYYYMMDD_HHMMSS' format
current_time_suffix= datetime.now().strftime("%Y%m%d_%H%M%S")

# Base name for the output file
base_output_filename = 'P04b_Customer_mastering_Fuzzy_output'


#base name for the output file
output_filename= f"{base_output_filename}_All_Sources_Fuzzy__Final_RESULT_{current_time_suffix}.xlsx"


# Function to calculate similarity
def similarity_score(str1, str2):
    return fuzz.ratio(str1.lower(), str2.lower()) / 100.0

# Load the dataset
df = pd.read_excel(latest_Fuzzy_Result_file, sheet_name='Customer_mastering_analysisPair')

set_link_score_threshold = 0.520 #This is the threshold for the link score. This has to be manually set upon inspection of the link score, and name and address matches.

# Filter pairs with Link Score < 0.605
filtered_df = df[df['Link Score'] < set_link_score_threshold]

# Group by 'Cluster ID' and aggregate data
grouped_df = filtered_df.groupby('Cluster ID').agg(list).reset_index()

# Define a function to process each group and assign labels, including similarity scores
def assign_labels(row):
    names = row['Name']
    addresses = [f"{addr}, {city}, {state} {str(zip)}" for addr, city, state, zip in zip(row['Address'], row['City'], row['State'], row['Zip'])]

    # Calculate similarity scores
    name_similarity = similarity_score(names[0], names[1])
    address_similarity = similarity_score(addresses[0], addresses[1])

    # Determine label
    if name_similarity >= 0.70 and address_similarity >= 0.80:
        label = 'Reliable match'
    elif name_similarity < 0.70 and address_similarity >= 0.80:
        label = 'Name mismatch, address same'
    elif name_similarity >= 0.70 and address_similarity < 0.80:
        label = 'Name match, address different'
    elif name_similarity < 0.70 and address_similarity < 0.80:
        label = 'Name and Address do not match'
    else:
        label = 'Other'

    return pd.Series([label, name_similarity, address_similarity], index=['Label', 'Name Similarity', 'Address Similarity'])

# Apply the function to each group and create a DataFrame with labels and similarity scores
labels_scores_df = grouped_df.apply(assign_labels, axis=1)
labels_scores_df['Cluster ID'] = grouped_df['Cluster ID']  # Add 'Cluster ID' to the DataFrame
print("Labels and Scores DataFrame Columns:", labels_scores_df.columns)  # Debugging

# Merge the labels and scores back into the original DataFrame
print("Original DataFrame Columns:", df.columns)  # Debugging
merged_df = df.merge(labels_scores_df, on='Cluster ID', how='left')

# Define a function to adjust 'Link Score' based on the label
def adjust_link_score(row):
    if row['Label'] in ['Name mismatch, address same']:
        return 0.3  # Replace with your chosen value below threshold, these are somewhat arbitrary numeric flags but need to be below the manually define threshold value to sort properly in the output file.
    elif row['Label'] in ['Name match, address different']:
        return 0.2  # Replace with your chosen value below threshold, these are somewhat arbitrary numeric flags but need to be below the manually define threshold value to sort properly in the output file.
    elif row['Label'] in ['Name and Address do not match']:
        return 0.1  # Replace with your chosen value below threshold, these are somewhat arbitrary numeric flags but need to be below the manually define threshold value to sort properly in the output file.
    elif row['Label'] in ['Reliable match']:
        return set_link_score_threshold
    else:
        return row['Link Score']

# Apply the function to adjust 'Link Score'
merged_df['Adjusted Link Score'] = merged_df.apply(adjust_link_score, axis=1)

# Fill NaN values in 'Label' column with 'Other' for rows with Link Score >= 0.605
merged_df['Label'].fillna('Other', inplace=True)



# First, create a helper column to distinguish between the two entries of each 'Cluster ID'
merged_df['Entry'] = merged_df.groupby('Cluster ID').cumcount() + 1

# Pivot the DataFrame
#Use the 'Entry' column as the columns parameter to pivot the DataFrame. Entry is the helper column created above, which distinguishes between the two entries of each 'Cluster ID'
#Use try accept to handle cases where one column is missing from the DataFrame, such as ShipTo Number which may not be present in the data. In such cases, try the pivot with the missing column and if it fails, pivot without it.
# try to generalize to handle any missing column in the values parameter
try:
    pivot_df = merged_df.pivot(index='Cluster ID', columns='Entry', values=['Id', 'Name', 'Address', 'City', 'State', 'Zip', 'Q_ID', 'ShipTo Number', 'Source', 'source file', 'Link Score', 'Adjusted Link Score', 'Label'])
except KeyError as e:
    missing_column = str(e).split("'")[1] # Get the missing column name from the error message
    print(f"Warning Missing column: {missing_column}") # Debugging
    pivot_df = merged_df.pivot(index='Cluster ID', columns='Entry', values=[col for col in merged_df.columns if col != missing_column]) # Pivot without the missing column
# Flatten the MultiIndex columns
pivot_df.columns = ['_'.join(str(col) for col in column) for column in pivot_df.columns]

# Rename columns based on suffix
new_column_names = []
for col in pivot_df.columns:
    if col.endswith('_1'):
        new_column_names.append('SFDC_Acc_' + col[:-2])  # Remove '_1' and add 'SFDC_Acc_'
    elif col.endswith('_2'):
        new_column_names.append('ExFactory_' + col[:-2])  # Remove '_2' and add 'ExFactory_'
    else:
        new_column_names.append(col)

pivot_df.columns = new_column_names


# Rename columns based on suffix and sort them
sfdc_columns = [col for col in pivot_df.columns if col.startswith('SFDC_Acc_')]
exfactory_columns = [col for col in pivot_df.columns if col.startswith('ExFactory_')]

# Sort the columns within each group
sfdc_columns.sort()
exfactory_columns.sort()

# Combine the sorted columns and add any remaining columns
sorted_columns = sfdc_columns + exfactory_columns + [col for col in pivot_df.columns if col not in sfdc_columns + exfactory_columns]

# Reorder the DataFrame columns
pivot_df = pivot_df[sorted_columns]

# Reset index to make 'Cluster ID' a column again
pivot_df.reset_index(inplace=True)

# Merge with the DataFrame containing Adjusted Link Score and Label
final_df = pivot_df.merge(merged_df[['Cluster ID', 'Adjusted Link Score', 'Label', 'Name Similarity', 'Address Similarity']].drop_duplicates(), on='Cluster ID')

# Define the desired column order
column_order = [
    'Adjusted Link Score', 'Cluster ID', 
    'SFDC_Acc_Id', 'SFDC_Acc_Name', 'SFDC_Acc_Address', 'SFDC_Acc_City', 
    'SFDC_Acc_State', 'SFDC_Acc_Zip', 'SFDC_Acc_Q_ID', 'SFDC_Acc_ShipTo Number', 
    'SFDC_Acc_Source', 'SFDC_Acc_source file', 
    'ExFactory_Address', 'ExFactory_City', 'ExFactory_Id', 'ExFactory_Link Score', 
    'ExFactory_Name', 'ExFactory_Q_ID', 'ExFactory_ShipTo Number', 
    'ExFactory_Source', 'ExFactory_State', 'ExFactory_Zip', 'ExFactory_source file', 
    'Address Similarity', 'Name Similarity', 'SFDC_Acc_Link Score', 'Label'
]

existing_columns = [col for col in column_order if col in final_df.columns]

#Print warning about any missing columns    
missing_columns = set(column_order) - set(existing_columns)
if missing_columns:
    print(f"WARNING: Missing columns if these are required for analysis of results consider checking input file for these columns: {missing_columns}")

# Reorder the DataFrame columns
final_df = final_df[existing_columns]


#Drop any columns with all NaN values .This will drop columns that contain all blank values. This is useful for removing unneeded columns that are filled with NaN (missing) values.
final_df.dropna(axis=1, how='all', inplace=True) # Drop any columns with all NaN values, axis=1 means columns, how='all' means all values in the column are NaN


# Construct the full path for the output file
output_file_path = os.path.join(base_path, output_filename)

# Save the results
final_df.to_excel(output_file_path, index=False)

###==================================Locate Mastering Analysis file needed for adding results to a new to EXCEL
# Define the base directory and the file prefix
base_directory = r"C:\Users\beste\OneDrive - Qral Group\01 Narcan\08 Mastering"
file_prefix = 'Mastering All Sources'

# Use glob to find all files that match the prefix in all subdirectories of the base directory
files = glob.glob(os.path.join(base_directory, '**', file_prefix + '*.xlsx'), recursive=True)

# If no files were found, print an error message and exit
if not files:
    print(f"No files found with prefix '{file_prefix}' in directory '{base_directory}'")
    exit(1)

# Define a function to extract the version number from a file name
def get_version_number(filename):
    match = re.search(r'v(\d+\.\d+)', filename)
    if match:
        return float(match.group(1))
    else:
        return 0

# Sort the files by version number, and select the one with the highest version number
latest_Mastering_All_Sources_Analysis_file = max(files, key=get_version_number)

print(f"Using file '{latest_Mastering_All_Sources_Analysis_file}'") # this will print the full path of the most recent file that will be used for checking new accounts against the logbook

# Get the parent directory of the latest file
input2_base_path = os.path.dirname(latest_Mastering_All_Sources_Analysis_file)

print(f"Using base path '{input2_base_path}'") # this will print the base path that will be used for loading files and saving the results of the analysis of new accounts

#latest_file = os.path.join(input2_base_path, 'Mastering Full Extract 20231218_114447.xlsx')

mastering_extract_filename = os.path.basename(latest_Mastering_All_Sources_Analysis_file) # use this to print the name of the file that will be used for checking new accounts against the logbook

# Assume 'final_df' is your DataFrame

# Load the existing workbook
book = load_workbook(latest_Mastering_All_Sources_Analysis_file)

# Create a new sheet in the workbook for 'final_df'
new_sheet = book.create_sheet('2a. New Fuzzy Results')

# Write the column headers to the new sheet
new_sheet.append(final_df.columns.tolist())

# Write the DataFrame to the new sheet
for r in final_df.to_dict(orient='records'):  # Convert DataFrame to list of dicts
    new_sheet.append(list(r.values()))  # Convert each dict to a list of values

# Save the workbook
book.save(latest_Mastering_All_Sources_Analysis_file)

print(f"Added new sheet to '{latest_Mastering_All_Sources_Analysis_file}'")